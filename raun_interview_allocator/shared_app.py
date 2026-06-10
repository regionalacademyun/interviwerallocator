import io
from datetime import date, timedelta

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from . import user_config as cfg
from .dashboard_components import app_styles, render_dashboard
from .data_io import load_applicants, make_streamlit_safe_df, to_excel_bytes
from .google_sheets_io import read_google_sheet, write_google_sheet
from .interview_logic import (
    allocate_interviews,
    build_default_interviewers_df,
    compute_planning_metrics,
    filter_interview_pool,
    normalize_candidate_availability,
    normalize_interviewers_input,
    parse_interview_source,
)


# =========================================================
# GENERAL UI HELPERS
# =========================================================
def get_logo_path():
    for p in cfg.LOGO_CANDIDATES:
        if p.exists():
            return p
    return None


def login_screen():
    app_styles()
    left, right = st.columns([1, 1.4])
    with left:
        logo = get_logo_path()
        if logo:
            st.image(str(logo), width=120)
        st.markdown('<div class="hero">', unsafe_allow_html=True)
        st.title(cfg.APP_TITLE)
        st.caption(cfg.APP_SUBTITLE)
        st.markdown(
            "<div class='explain'>Choose your name and enter the password to open the app. "
            "If your name is not listed, choose <b>New user / guest reviewer</b> and type your name.</div>",
            unsafe_allow_html=True,
        )
        st.markdown('</div>', unsafe_allow_html=True)
    with right:
        st.markdown('<div class="step">', unsafe_allow_html=True)
        login_options = [""] + cfg.RAUN_TEAM_MEMBERS + [cfg.NEW_USER_LABEL]
        selected_user = st.selectbox("Choose your name", login_options, index=0)
        custom_user = ""
        if selected_user == cfg.NEW_USER_LABEL:
            custom_user = st.text_input("Type your full name")
        password = st.text_input("Password", type="password")
        if st.button("Login", type="primary", width="stretch"):
            final_user = custom_user.strip() if selected_user == cfg.NEW_USER_LABEL else selected_user
            if not final_user:
                st.warning("Please choose or type your name.")
            elif password == cfg.APP_PASSWORD:
                st.session_state.logged_in = True
                st.session_state.username = final_user
                st.rerun()
            else:
                st.error("Wrong password")
        st.markdown('</div>', unsafe_allow_html=True)
        st.stop()


def glossy_header(username: str):
    st.markdown(
        f"""
        <div class="hero">
            <h2 style="margin-bottom:0.25rem;">Hello {username.split()[0] if username else 'there'}!</h2>
            <div class="explain">This is the <b>interview allocation app</b>. It reads the RAUN interview-allocation sheet, derives the interview pool from Reviewer 1 and Reviewer 2 decisions, generates clean availability templates, and then matches candidates and interviewers by shared date/time while favouring one of the two original reviewers wherever possible.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def ensure_state():
    defaults = {
        "source_mode": "Offline upload",
        "connected_sheet_ref": "",
        "connected_worksheet": "Sheet1",
        "local_service_account_file": "",
        "raw_input_df": None,
        "applicants_df": None,
        "review_source_df": None,
        "interview_pool_df": None,
        "interviewers_working_df": None,
        "interviewer_source_label": "Uniform RAUN interview baseline table",
        "candidate_availability_df": None,
        "candidate_availability_source_label": "No candidate availability uploaded",
        "intalloc_df": None,
        "intloads_df": None,
        "intexceptions_df": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def read_uploaded_raw(uploaded_file):
    if uploaded_file is None:
        return None
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file, header=None)
    return pd.read_excel(uploaded_file, sheet_name=0, header=None)


# =========================================================
# AVAILABILITY TEMPLATE GENERATION
# =========================================================
def _build_slots(start: date, end: date, time_values: list[str]) -> list[tuple[str, str]]:
    slots = []
    if end < start:
        start, end = end, start
    cur = start
    while cur <= end:
        for t in time_values:
            slots.append((cur.strftime("%d-%m-%Y"), str(t)))
        cur += timedelta(days=1)
    return slots


def _write_availability_template(sheet_title: str, id_headers: list[str], rows: list[list], slots: list[tuple[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    blue = PatternFill("solid", fgColor="D9EAF7")
    green = PatternFill("solid", fgColor="D9EAD3")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    grey = PatternFill("solid", fgColor="F3F4F6")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Identity columns occupy the first len(id_headers) columns.
    n_id = len(id_headers)
    for c, h in enumerate(id_headers, start=1):
        ws.cell(1, c).value = h
        ws.cell(2, c).value = ""
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(1, c).fill = grey
        ws.cell(2, c).fill = grey
        ws.cell(1, c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(1, c).border = border
        ws.cell(2, c).border = border

    for idx, (d, t) in enumerate(slots, start=n_id + 1):
        ws.cell(1, idx).value = d
        ws.cell(2, idx).value = t
        ws.cell(1, idx).font = Font(bold=True)
        ws.cell(2, idx).font = Font(bold=True)
        ws.cell(1, idx).fill = blue
        ws.cell(2, idx).fill = yellow
        ws.cell(1, idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(1, idx).border = border
        ws.cell(2, idx).border = border

    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = val
            ws.cell(r_idx, c_idx).border = border
            ws.cell(r_idx, c_idx).alignment = Alignment(vertical="center")
        for c_idx in range(n_id + 1, n_id + len(slots) + 1):
            cell = ws.cell(r_idx, c_idx)
            cell.value = "No"
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dropdown for availability cells.
    if slots and rows:
        first = f"{get_column_letter(n_id + 1)}3"
        last = f"{get_column_letter(n_id + len(slots))}{len(rows) + 2}"
        dv = DataValidation(type="list", formula1='"Yes,No,Under reserve"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{first}:{last}")

    ws.freeze_panes = ws.cell(3, n_id + 1).coordinate
    for c_idx in range(1, n_id + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 22 if c_idx != 3 else 34
    for c_idx in range(n_id + 1, n_id + len(slots) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    # Add an instructions sheet.
    ins = wb.create_sheet("Instructions")
    ins["A1"] = "RAUN availability template"
    ins["A1"].font = Font(bold=True, size=14)
    ins["A3"] = "Fill only the date/time cells with Yes, No, or Under reserve."
    ins["A4"] = "The app reads row 1 as dates, row 2 as times, and row 3 onwards as names."
    ins["A5"] = "Do not move, delete, or rename the identity columns."
    ins["A6"] = "Yes means available. Under reserve is kept for admin reference but is not treated as a confirmed available slot."
    ins.column_dimensions["A"].width = 110

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def make_interviewer_availability_template_bytes(interviewers_df: pd.DataFrame, slots: list[tuple[str, str]]) -> bytes:
    if interviewers_df is None or interviewers_df.empty:
        rows = [[name] for name in cfg.RAUN_TEAM_MEMBERS]
    else:
        names = interviewers_df["Reviewer Name"].fillna("").astype(str).str.strip().tolist() if "Reviewer Name" in interviewers_df.columns else cfg.RAUN_TEAM_MEMBERS
        rows = [[name] for name in names if name]
    return _write_availability_template("Interviewer Availability", ["Interviewer"], rows, slots)


def make_candidate_availability_template_bytes(interview_pool_df: pd.DataFrame, slots: list[tuple[str, str]]) -> bytes:
    rows = []
    if interview_pool_df is not None and not interview_pool_df.empty:
        for _, r in interview_pool_df.iterrows():
            rows.append([
                int(r.get("Applicant ID", len(rows) + 1)) if str(r.get("Applicant ID", "")).strip() else len(rows) + 1,
                str(r.get("Full Name", "")).strip(),
                str(r.get("Email", "")).strip(),
            ])
    return _write_availability_template("Candidate Availability", ["Applicant ID", "Candidate", "Email"], rows, slots)


def make_interviewer_template_bytes(interview_pool_df):
    template_df = build_default_interviewers_df(
        cfg.RAUN_TEAM_MEMBERS,
        interview_pool_df=interview_pool_df,
        use_uniform_baseline=True,
    )
    return to_excel_bytes({"Interviewer Baseline Input": template_df})


# =========================================================
# PLOTS / PARSING
# =========================================================
def make_interview_capacity_plot(loads_df: pd.DataFrame):
    if loads_df is None or loads_df.empty:
        return None
    df = loads_df.copy().sort_values("Interview Assigned", ascending=False)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df["Reviewer Name"],
        y=df["Interview Capacity"],
        name="Available slot capacity",
        marker_color="#BFDBFE",
        hovertemplate="<b>%{x}</b><br>Capacity: %{y}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=df["Reviewer Name"],
        y=df["Interview Assigned"],
        mode="markers+lines",
        name="Actually assigned",
        marker=dict(size=12, color="#8B5CF6", symbol="diamond"),
        line=dict(color="#8B5CF6", width=3),
        customdata=df[["Interview Capacity", "Remaining Interview Capacity", "Interview Utilization %"]].values,
        hovertemplate=(
            "<b>%{x}</b><br>Assigned: %{y}<br>Capacity: %{customdata[0]}<br>Remaining: %{customdata[1]}<br>Utilization: %{customdata[2]}%<extra></extra>"
        ),
    ))
    fig.update_layout(
        height=470,
        template="plotly_white",
        title="Interviewer capacity check: available slots and actual assignments",
        xaxis_title="Interviewer",
        yaxis_title="Number of interviews",
        margin=dict(l=20, r=20, t=70, b=20),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
    )
    return fig


def _load_review_source_from_current_raw(decision_policy):
    raw = st.session_state.raw_input_df
    if raw is None or raw.empty:
        st.warning("Load the main Excel/Google Sheet first.")
        return
    parsed = parse_interview_source(raw, decision_policy=decision_policy)
    st.session_state.review_source_df = parsed
    st.session_state.interview_pool_df = filter_interview_pool(parsed)
    st.success(f"Review decisions parsed: {len(parsed)} applicants found; {len(st.session_state.interview_pool_df)} marked for interview.")


def _read_availability_upload(uploaded_file):
    if uploaded_file is None:
        return None
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file, header=None)
    return pd.read_excel(uploaded_file, sheet_name=0, header=None)


# =========================================================
# MAIN APP
# =========================================================
def main_app():
    app_styles()
    ensure_state()
    glossy_header(st.session_state.get("username", ""))

    st.markdown('<div class="step">', unsafe_allow_html=True)
    st.markdown("## Step 1 — Choose data source and load the interview-allocation sheet")
    st.markdown("<div class='explain'>Use the Excel/Google Sheet where data starts from row 3 and includes Reviewer 1, Reviewer 2, applicant details, both reviewer scoring blocks, and the two <b>Invitation to interview?</b> columns.</div>", unsafe_allow_html=True)

    source_mode = st.radio(
        "Input data source",
        ["Offline upload", "Connected Google Sheet"],
        horizontal=True,
        index=0 if st.session_state.source_mode == "Offline upload" else 1,
        help="Offline uses an uploaded Excel/CSV. Connected mode reads a live Google Sheet only after you paste the sheet address/ID.",
    )
    st.session_state.source_mode = source_mode

    if source_mode == "Offline upload":
        applicant_file = st.file_uploader("Upload interview-allocation Excel or CSV", type=["xlsx", "csv"], help="This is the sheet containing applicant details and Reviewer 1/Reviewer 2 preselection assessment columns.")
        if applicant_file is not None:
            try:
                raw_df = read_uploaded_raw(applicant_file)
                st.session_state.raw_input_df = raw_df
                st.session_state.applicants_df = load_applicants(raw_df)
                st.success(f"Input file loaded: {len(st.session_state.applicants_df)} applicants detected for dashboard view.")
            except Exception as e:
                st.error(f"Could not read input file: {e}")
                st.markdown('</div>', unsafe_allow_html=True)
                return
    else:
        c1, c2 = st.columns([1.6, 1])
        with c1:
            sheet_ref = st.text_input("Google Sheet address or ID", value=st.session_state.connected_sheet_ref)
        with c2:
            worksheet = st.text_input("Worksheet/tab name", value=st.session_state.connected_worksheet)
        service_path = st.text_input("Optional local JSON key path for testing only", value=st.session_state.local_service_account_file, type="password")
        st.session_state.connected_sheet_ref = sheet_ref
        st.session_state.connected_worksheet = worksheet
        st.session_state.local_service_account_file = service_path
        if st.button("Load live Google Sheet", type="primary", width="stretch"):
            try:
                raw_df = read_google_sheet(sheet_ref, worksheet, service_account_file=service_path or None)
                st.session_state.raw_input_df = raw_df
                st.session_state.applicants_df = load_applicants(raw_df)
                st.success(f"Live Google Sheet loaded: {len(st.session_state.applicants_df)} applicants detected for dashboard view.")
            except Exception as e:
                st.error(f"Could not read Google Sheet: {e}")
                st.markdown('</div>', unsafe_allow_html=True)
                return

    applicants_df = st.session_state.applicants_df
    st.markdown('</div>', unsafe_allow_html=True)
    if applicants_df is None or applicants_df.empty:
        st.info("Please load the interview-allocation sheet to continue.")
        return

    preview_reviewers = st.session_state.interviewers_working_df
    if preview_reviewers is None:
        preview_reviewers = build_default_interviewers_df(cfg.RAUN_TEAM_MEMBERS, interview_pool_df=pd.DataFrame(), use_uniform_baseline=False)
    render_dashboard(applicants_df, preview_reviewers.copy(), shortlist_only=True)

    st.markdown('<div class="step">', unsafe_allow_html=True)
    st.markdown("## Step 2 — Read Reviewer 1/2 assessment decisions")
    st.markdown("<div class='explain'>The app reads the two <b>Invitation to interview?</b> columns and applies the RAUN logic. The safe default is: <b>include everyone unless the combined decision is explicitly No</b>.</div>", unsafe_allow_html=True)
    decision_policy = st.selectbox(
        "Who should enter the interview allocation pool?",
        ["Exclude only explicit No+No", "Exclude No and No decision yet", "Yes + Maybe+Yes", "Yes only", "Yes + Maybe+Yes + Maybe"],
        index=0,
        help="Recommended default: include everyone unless both reviewer decisions produce an explicit No.",
    )
    if st.button("Parse interview decisions from loaded sheet", type="primary", width="stretch"):
        _load_review_source_from_current_raw(decision_policy)

    if st.session_state.review_source_df is not None:
        all_review_df = st.session_state.review_source_df
        pool_df = st.session_state.interview_pool_df
        a, b, c, d = st.columns(4)
        a.metric("Applicants parsed", len(all_review_df))
        b.metric("Interview pool", len(pool_df))
        c.metric("Not included", max(0, len(all_review_df) - len(pool_df)))
        d.metric("Decision policy", decision_policy)
        st.markdown("### Decision breakdown")
        if "Final Result- First review" in all_review_df.columns:
            breakdown = all_review_df["Final Result- First review"].fillna("No decision yet").astype(str).str.strip().replace({"": "No decision yet"}).value_counts().reset_index()
            breakdown.columns = ["First-review result", "Candidates"]
            st.dataframe(make_streamlit_safe_df(breakdown), width="stretch", height=220)
        st.markdown("### Reviewer coverage")
        reviewer_cols = [c for c in ["Preselection Reviewer 1", "Preselection Reviewer 2"] if c in all_review_df.columns]
        if reviewer_cols:
            reviewers_long = all_review_df[reviewer_cols].melt(value_name="Reviewer")["Reviewer"].fillna("").astype(str).str.strip()
            reviewers_long = reviewers_long[reviewers_long.ne("")]
            reviewer_summary = reviewers_long.value_counts().reset_index()
            reviewer_summary.columns = ["Reviewer", "Assigned preselection reviews in sheet"]
            st.dataframe(make_streamlit_safe_df(reviewer_summary), width="stretch", height=260)
        st.markdown("### Parsed decision preview")
        preview_cols = [c for c in ["Applicant ID", "Full Name", "Preselection Reviewer 1", "Preselection Reviewer 2", "Reviewer 1 Invitation", "Reviewer 2 Invitation", "Final Result- First review", "Interview Required", "Total points from Reviewer 1", "Total points from Reviewer 2", "Total points from both reviewers"] if c in all_review_df.columns]
        st.dataframe(make_streamlit_safe_df(all_review_df[preview_cols]), width="stretch", height=300)
        st.markdown("### Interview pool only")
        st.dataframe(make_streamlit_safe_df(pool_df[preview_cols]), width="stretch", height=300)
    st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.interview_pool_df is None or st.session_state.interview_pool_df.empty:
        st.info("Parse interview decisions first. The availability and matching controls will appear after the interview pool is found.")
        return

    interview_pool_df = st.session_state.interview_pool_df
    if st.session_state.interviewers_working_df is None:
        st.session_state.interviewers_working_df = build_default_interviewers_df(cfg.RAUN_TEAM_MEMBERS, interview_pool_df=interview_pool_df, use_uniform_baseline=True)
        st.session_state.interviewer_source_label = "Uniform RAUN interview baseline table"

    # =========================================================
    # STEP 3: GENERATE + LOAD AVAILABILITY
    # =========================================================
    st.markdown('<div class="step">', unsafe_allow_html=True)
    st.markdown("## Step 3 — Generate and load availability files")
    st.markdown("<div class='explain'>First generate clean interviewer and candidate availability templates using your chosen dates and times. Send them out, collect the filled files, then upload both completed templates here. The app then matches candidates and interviewers using only shared date/time slots.</div>", unsafe_allow_html=True)

    st.markdown("### 3A — Generate clean availability templates")
    g1, g2, g3 = st.columns([1, 1, 2])
    with g1:
        start_date = st.date_input("First interview date", value=date.today(), key="avail_start_date")
    with g2:
        end_date = st.date_input("Last interview date", value=date.today() + timedelta(days=4), key="avail_end_date")
    with g3:
        all_times = [f"{h:02d}:00" for h in range(8, 20)]
        default_times = ["09:00", "10:00", "11:00", "12:00", "13:00", "14:00", "15:00", "16:00", "17:00"]
        selected_times = st.multiselect("Interview time slots", all_times, default=default_times, key="avail_times")
    slots = _build_slots(start_date, end_date, selected_times)
    st.caption(f"Template will contain {len(slots)} date/time slots.")

    baseline_interviewers = st.session_state.interviewers_working_df.copy()
    int_template = make_interviewer_availability_template_bytes(baseline_interviewers, slots)
    cand_template = make_candidate_availability_template_bytes(interview_pool_df, slots)
    t1, t2 = st.columns(2)
    with t1:
        st.download_button(
            "Download interviewer availability template",
            data=int_template,
            file_name="RAUN_interviewer_availability_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    with t2:
        st.download_button(
            "Download candidate availability template",
            data=cand_template,
            file_name="RAUN_candidate_availability_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

    st.markdown("### 3B — Upload completed availability files")
    u1, u2 = st.columns(2)
    with u1:
        interviewer_file = st.file_uploader("Upload completed interviewer availability", type=["xlsx", "csv"], key="int_upload")
        if interviewer_file is not None:
            try:
                raw_int = _read_availability_upload(interviewer_file)
                st.session_state.interviewers_working_df = normalize_interviewers_input(raw_int)
                st.session_state.interviewer_source_label = f"Loaded from file: {interviewer_file.name}"
                st.success("Interviewer availability loaded.")
            except Exception as e:
                st.error(f"Could not read interviewer availability: {e}")
    with u2:
        cand_file = st.file_uploader("Upload completed candidate availability", type=["xlsx", "csv"], key="candidate_availability")
        if cand_file is not None:
            try:
                raw_cand = _read_availability_upload(cand_file)
                st.session_state.candidate_availability_df = normalize_candidate_availability(raw_cand)
                st.session_state.candidate_availability_source_label = f"Loaded from file: {cand_file.name}"
                st.success("Candidate availability loaded.")
            except Exception as e:
                st.error(f"Could not read candidate availability: {e}")

    st.info(f"Current interviewer table source: {st.session_state.interviewer_source_label}")
    interviewers_df = st.session_state.interviewers_working_df.copy()
    st.markdown("#### Parsed interviewer availability")
    st.dataframe(make_streamlit_safe_df(interviewers_df), width="stretch", height=260)

    if st.session_state.candidate_availability_df is not None:
        st.info(f"Current candidate availability source: {st.session_state.candidate_availability_source_label}")
        st.markdown("#### Parsed candidate availability")
        st.dataframe(make_streamlit_safe_df(st.session_state.candidate_availability_df), width="stretch", height=260)
    else:
        st.warning("Candidate availability has not been uploaded yet. For final matching, upload the completed candidate availability template.")
    st.markdown('</div>', unsafe_allow_html=True)

    # =========================================================
    # STEP 4: GENERATE MATCHING
    # =========================================================
    st.markdown('<div class="step">', unsafe_allow_html=True)
    st.markdown("## Step 4 — Generate final interview matching")
    st.markdown("<div class='explain'>The app first tries to match each candidate with one of the two original reviewers, but only if they share a date/time slot. If that is not possible, it matches the candidate with another available team member. If no shared slot exists, the candidate is left unmatched and flagged.</div>", unsafe_allow_html=True)

    seed = st.number_input("Random seed", min_value=1, value=42, step=1, help="Use the same seed to reproduce the same allocation.")
    match_strength = st.selectbox("Background matching strength", ["Off", "Low", "Medium"], index=1, help="Soft bonus only. Continuity and availability remain more important.")
    prefer_continuity = st.checkbox("Prioritise one of the two original preselection reviewers", value=True)
    require_overlap = st.checkbox(
        "Require candidate/interviewer availability overlap",
        value=True,
        help="Recommended ON for final matching. When ON, the app only assigns interviews where candidate and interviewer share the same date/time slot.",
    )

    planning = compute_planning_metrics(interview_pool_df, interviewers_df)
    p1, p2, p3, p4 = st.columns(4)
    p1.metric("Interview candidates", planning["Interview Candidates"])
    p2.metric("Available interviewer slots", planning["Interview Capacity"])
    p3.metric("Capacity gap", planning["Capacity Gap"])
    p4.metric("Shortage", planning["Shortage"])

    tab_alloc, tab_stats, tab_sync = st.tabs(["Final Matching", "Workload Stats", "Google Sheet Output"])
    with tab_alloc:
        if st.button("Generate final interview matching", type="primary", width="stretch"):
            intalloc_df, intloads_df, intexceptions_df = allocate_interviews(
                interview_pool_df=interview_pool_df,
                interviewers_df=interviewers_df,
                candidate_availability_df=st.session_state.candidate_availability_df,
                seed=int(seed),
                match_strength=match_strength,
                prefer_preselection_reviewer=prefer_continuity,
                require_availability_overlap=require_overlap,
            )
            st.session_state.intalloc_df = intalloc_df
            st.session_state.intloads_df = intloads_df
            st.session_state.intexceptions_df = intexceptions_df
        if st.session_state.intalloc_df is not None:
            final_cols = [c for c in ["Interviewer", "Candidate", "Date", "Time", "Email address", "Interview email sent", "Interviewee confirmed"] if c in st.session_state.intalloc_df.columns]
            final_matching_df = st.session_state.intalloc_df[final_cols].copy() if final_cols else st.session_state.intalloc_df.copy()
            st.markdown("### Final matching output")
            st.dataframe(make_streamlit_safe_df(final_matching_df), width="stretch", height=420)
            st.markdown("### Matching detail")
            st.dataframe(make_streamlit_safe_df(st.session_state.intalloc_df), width="stretch", height=320)
            st.markdown("### Exceptions / unmatched")
            if st.session_state.intexceptions_df is not None and not st.session_state.intexceptions_df.empty:
                st.dataframe(make_streamlit_safe_df(st.session_state.intexceptions_df), width="stretch", height=240)
            else:
                st.success("No interview exceptions.")
            out_excel = to_excel_bytes({
                "Final Matching": final_matching_df,
                "Interview Allocation Detail": st.session_state.intalloc_df,
                "Interviewer Loads": st.session_state.intloads_df,
                "Interview Exceptions": st.session_state.intexceptions_df,
                "Interview Pool": interview_pool_df,
                "Parsed Review Decisions": st.session_state.review_source_df,
                "Parsed Candidate Availability": st.session_state.candidate_availability_df,
                "Parsed Interviewer Availability": interviewers_df,
            })
            st.download_button("Download full interview matching Excel", out_excel, "raun_final_interview_matching.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch")
    with tab_stats:
        if st.session_state.intloads_df is None:
            st.info("Generate final matching first to see workload stats.")
        else:
            st.dataframe(make_streamlit_safe_df(st.session_state.intloads_df), width="stretch", height=300)
            fig = make_interview_capacity_plot(st.session_state.intloads_df)
            if fig is not None:
                st.plotly_chart(fig, width="stretch", key="interview_capacity_chart")
    with tab_sync:
        st.subheader("Write interview matching results to Google Sheet")
        st.caption("Optional. Nothing is written unless you press a write button.")
        if st.session_state.intalloc_df is None:
            st.info("Generate final matching first.")
        else:
            final_cols = [c for c in ["Interviewer", "Candidate", "Date", "Time", "Email address", "Interview email sent", "Interviewee confirmed"] if c in st.session_state.intalloc_df.columns]
            final_matching_df = st.session_state.intalloc_df[final_cols].copy() if final_cols else st.session_state.intalloc_df.copy()
            out_sheet_ref = st.text_input("Output Google Sheet address or ID", value=st.session_state.connected_sheet_ref)
            service_path_out = st.text_input("Optional local JSON key path", value=st.session_state.local_service_account_file, type="password", key="output_service_path")
            c1, c2, c3 = st.columns(3)
            with c1:
                if st.button("Write final matching tab", width="stretch"):
                    try:
                        write_google_sheet(final_matching_df, out_sheet_ref, "Final Matching", service_account_file=service_path_out or None)
                        st.success("Final Matching written.")
                    except Exception as e:
                        st.error(f"Could not write final matching: {e}")
            with c2:
                if st.button("Write loads tab", width="stretch"):
                    try:
                        write_google_sheet(st.session_state.intloads_df, out_sheet_ref, "Interviewer Loads", service_account_file=service_path_out or None)
                        st.success("Interviewer Loads written.")
                    except Exception as e:
                        st.error(f"Could not write loads: {e}")
            with c3:
                if st.button("Write exceptions tab", width="stretch"):
                    try:
                        write_google_sheet(st.session_state.intexceptions_df, out_sheet_ref, "Interview Exceptions", service_account_file=service_path_out or None)
                        st.success("Interview Exceptions written.")
                    except Exception as e:
                        st.error(f"Could not write exceptions: {e}")
    st.markdown('</div>', unsafe_allow_html=True)
